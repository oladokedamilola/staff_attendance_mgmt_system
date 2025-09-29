import csv

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib import messages

User = get_user_model()

from .models import Leave
from .forms import LeaveApplicationForm

from .utils import send_leave_notification
@login_required
def apply_leave(request):
    """Staff can apply for leave. Admins are notified via email and on-site notifications."""
    if not request.user.is_staff_user():
        messages.error(request, "Access denied. Staff account required.")
        return redirect("login")

    if request.method == "POST":
        form = LeaveApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.staff = request.user
            leave.save()

            try:
                # Notify all admins
                admins = User.objects.filter(role="admin")
                if admins.exists():
                    send_leave_notification(
                        sender=request.user,
                        recipient=admins,
                        subject="New Leave Request Submitted",
                        message=(
                            f"{request.user.get_full_name()} submitted a new leave request "
                            f"({leave.leave_type}) from {leave.start_date} to {leave.end_date}."
                        ),
                        request=request  # optional on-site notification
                    )
            except Exception as e:
                # Log error and notify staff
                messages.warning(
                    request,
                    f"Leave submitted, but failed to notify admins. Error logged."
                )

            messages.success(request, "Leave application submitted successfully.")
            return redirect("my_leave_requests")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = LeaveApplicationForm()

    return render(request, "staff/apply_leave.html", {"form": form})


import csv
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Leave
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

from django.db.models import Count
from django.db.models.functions import TruncDate

from django.db.models import Count
from django.db.models.functions import TruncDate
from django.contrib import messages
from django.shortcuts import render, redirect
from leave.models import Leave
import json

@login_required
def leave_report(request):
    if not request.user.is_admin_user():
        messages.error(request, "Access denied. Admin privileges required.")
        return redirect("login")

    leaves = Leave.objects.all().order_by("staff", "-applied_at")

    # Filters
    staff_email = request.GET.get("staff")
    status = request.GET.get("status")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if staff_email:
        leaves = leaves.filter(staff__email__icontains=staff_email)
    if status:
        leaves = leaves.filter(status__iexact=status.lower())
    if start_date:
        leaves = leaves.filter(start_date__gte=start_date)
    if end_date:
        leaves = leaves.filter(end_date__lte=end_date)

    if not leaves.exists():
        messages.warning(request, "No leave records found to display.")

    # Aggregate daily leave counts
    trend_data = leaves.annotate(date_only=TruncDate('applied_at')) \
                       .values('date_only', 'status') \
                       .annotate(count=Count('id')) \
                       .order_by('date_only')

    # Chart labels
    chart_labels = sorted({td['date_only'].strftime("%Y-%m-%d") for td in trend_data})

    # Initialize trends with 0 counts
    pending_trend = [0] * len(chart_labels)
    approved_trend = [0] * len(chart_labels)
    rejected_trend = [0] * len(chart_labels)

    # Fill in counts
    for td in trend_data:
        date_str = td['date_only'].strftime("%Y-%m-%d")
        index = chart_labels.index(date_str)
        if td['status'].lower() == 'pending':
            pending_trend[index] = td['count']
        elif td['status'].lower() == 'approved':
            approved_trend[index] = td['count']
        elif td['status'].lower() == 'rejected':
            rejected_trend[index] = td['count']

    # Total counts for badges
    pending_count = leaves.filter(status="pending").count()
    approved_count = leaves.filter(status="approved").count()
    rejected_count = leaves.filter(status="rejected").count()

    return render(request, "leave/admin_requests.html", {
        "leave_requests": leaves,
        "pending_count": pending_count,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
        # Convert to JSON for safe JS consumption
        "chart_labels": json.dumps(chart_labels),
        "pending_trend": json.dumps(pending_trend),
        "approved_trend": json.dumps(approved_trend),
        "rejected_trend": json.dumps(rejected_trend),
    })



@login_required
def leave_export(request, export_format):
    """Export leave report as CSV or Excel."""
    if not request.user.is_admin_user():
        messages.error(request, "Access denied. Admin privileges required.")
        return redirect("login")

    leaves = Leave.objects.all().order_by("staff", "-applied_at")

    # Apply the same filters as in leave_report
    staff_email = request.GET.get("staff")
    status = request.GET.get("status")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if staff_email:
        leaves = leaves.filter(staff__email__icontains=staff_email)
    if status:
        leaves = leaves.filter(status__iexact=status.lower())
    if start_date:
        leaves = leaves.filter(start_date__gte=start_date)
    if end_date:
        leaves = leaves.filter(end_date__lte=end_date)

    if export_format == "csv":
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="leave_report.csv"'

        writer = csv.writer(response)
        writer.writerow(["Staff", "Email", "Leave Type", "Start Date", "End Date", "Status", "Reason"])
        for leave in leaves:
            writer.writerow([
                leave.staff.get_full_name,
                leave.staff.email,
                leave.leave_type,
                leave.start_date,
                leave.end_date,
                leave.status.capitalize(),
                leave.reason
            ])
        return response

    elif export_format == "xlsx":
        import openpyxl
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leave Report"

        headers = ["Staff", "Email", "Leave Type", "Start Date", "End Date", "Status", "Reason"]
        ws.append(headers)

        for leave in leaves:
            ws.append([
                leave.staff.get_full_name,
                leave.staff.email,
                leave.leave_type,
                leave.start_date,
                leave.end_date,
                leave.status.capitalize(),
                leave.reason
            ])

        for i, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="leave_report.xlsx"'
        return response

    else:
        messages.error(request, "Invalid export format.")
        return redirect("leave_report")


from .models import Notification
from django.core.paginator import Paginator

@login_required
def notifications_list(request):
    notifications = Notification.objects.filter(recipient=request.user).order_by('-created_at')
    paginator = Paginator(notifications, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Check if "Mark all as read" should display
    show_mark_all = notifications.filter(is_read=False).count() > 6

    context = {
        "page_obj": page_obj,
        "show_mark_all": show_mark_all,
    }
    return render(request, "notifications/list.html", context)


@login_required
def mark_notification_read(request, notification_id):
    notification = Notification.objects.filter(id=notification_id, recipient=request.user).first()
    if notification and not notification.is_read:
        notification.is_read = True
        notification.save()
    return redirect('notifications_list')


@login_required
def mark_all_notifications_read(request):
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    return redirect('notifications_list')