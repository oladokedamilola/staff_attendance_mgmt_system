from datetime import date
import csv

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse

from .models import Attendance

from datetime import date, datetime, time
from django.utils.timezone import now

@login_required
def mark_attendance(request):
    """Allow staff to mark their own attendance (time-sensitive)."""
    if not request.user.is_staff_user():
        messages.error(request, "Access denied. Staff account required.")
        return redirect("login")

    today = date.today()
    attendance, created = Attendance.objects.get_or_create(staff=request.user, date=today)

    if request.method == "POST":
        current_time = now().time()
        cutoff = time(8, 10)  # 8:10 AM

        # Determine status based on current time
        if current_time <= cutoff:
            status = "present"
        else:
            status = "late"

        attendance.status = status
        attendance.timestamp = now()
        attendance.save()

        messages.success(request, f"Your attendance for today has been marked as '{attendance.status}'.")
        return redirect("staff_dashboard")

    else:
        if attendance.status:
            messages.info(request, f"Today's attendance already marked as '{attendance.status}'.")
        else:
            messages.info(request, "You haven't marked your attendance yet for today.")

    return render(request, "attendance/staff_mark.html", {"attendance": attendance})

import csv
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Attendance
from accounts.models import User
from django.db import models
from datetime import timedelta
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db import models
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
from accounts.models import User
from .models import Attendance

@login_required
def attendance_report(request):
    """Admins can view attendance reports with filtering + charts."""
    if not request.user.is_admin_user():
        messages.error(request, "Access denied. Admin privileges required.")
        return redirect("login")

    # Base queryset
    records = Attendance.objects.all().order_by("staff", "-date")

    # Filtering
    staff_email = request.GET.get("staff", "")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if staff_email:
        records = records.filter(staff__email=staff_email)
    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)

    # --- Status distribution ---
    status_counts = {
        "Present": records.filter(status="present").count(),
        "Absent": records.filter(status="absent").count(),
        "Late": records.filter(status="late").count(),
    }
    status_labels = list(status_counts.keys())
    status_data = list(status_counts.values())

    # --- Daily Trends ---
    # Collect all dates in range (if no records exist, generate a small range)
    if start_date and end_date:
        date_list = [
            start_date + timedelta(days=i)
            for i in range((datetime.strptime(end_date, "%Y-%m-%d") - datetime.strptime(start_date, "%Y-%m-%d")).days + 1)
        ]
    else:
        date_list = sorted({r.date for r in records})  # unique sorted dates from records

    # Create a dict keyed by date for easy lookup
    trend_dict = {}
    for d in date_list:
        trend_dict[str(d)] = {"present": 0, "absent": 0, "late": 0}

    # Fill counts from actual records
    daily_qs = (
        records.values("date")
        .annotate(
            present=models.Count("id", filter=models.Q(status="present")),
            absent=models.Count("id", filter=models.Q(status="absent")),
            late=models.Count("id", filter=models.Q(status="late")),
        )
    )

    for item in daily_qs:
        date_str = str(item["date"])
        trend_dict[date_str] = {
            "present": item["present"],
            "absent": item["absent"],
            "late": item["late"],
        }

    # Prepare final lists for Chart.js
    trend_labels = list(trend_dict.keys())
    trend_present = [trend_dict[d]["present"] for d in trend_labels]
    trend_absent = [trend_dict[d]["absent"] for d in trend_labels]
    trend_late = [trend_dict[d]["late"] for d in trend_labels]

    return render(
        request,
        "adminpanel/attendance_reports.html",
        {
            "attendance_records": records,
            "status_labels": status_labels,
            "status_data": status_data,
            "trend_labels": trend_labels,
            "trend_present": trend_present,
            "trend_absent": trend_absent,
            "trend_late": trend_late,
            "staff_list": User.objects.filter(role="staff"),
            "selected_staff": staff_email,
        },
    )



@login_required
def attendance_export(request, file_type):
    """Export attendance report (CSV or XLSX)."""
    if not request.user.is_admin_user():
        messages.error(request, "Access denied. Admin privileges required.")
        return redirect("login")

    # Base queryset
    records = Attendance.objects.all().order_by("staff", "-date")

    # Apply same filters
    staff_email = request.GET.get("staff")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if staff_email:
        records = records.filter(staff__email=staff_email)
    if start_date:
        records = records.filter(date__gte=start_date)
    if end_date:
        records = records.filter(date__lte=end_date)

    # --- CSV Export ---
    if file_type == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="attendance_report.csv"'
        writer = csv.writer(response)
        writer.writerow(["Staff", "Email", "Date", "Status"])
        for r in records:
            writer.writerow([r.staff.get_full_name(), r.staff.email, r.date, r.status])
        return response

    # --- Excel Export ---
    elif file_type == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        headers = ["Staff", "Email", "Date", "Status"]
        ws.append(headers)

        for r in records:
            ws.append([r.staff.get_full_name(), r.staff.email, str(r.date), r.status])

        # Auto adjust column width
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="attendance_report.xlsx"'
        wb.save(response)
        return response

    else:
        messages.error(request, "Invalid export type")
        return redirect("attendance_report")

